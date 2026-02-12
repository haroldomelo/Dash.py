import pandas as pd
from pyparsing import col
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import time
from datetime import date
import zipfile
import os
from pathlib import Path
from openpyxl import load_workbook

def setorB3():

    options = Options()
    options.headless = False
    diretorio = Path(__file__).parent

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    
    url = "https://www.b3.com.br/pt_br/produtos-e-servicos/negociacao/renda-variavel/empresas-listadas.htm"
    driver.get(url)

    WebDriverWait(driver, 10).until(
        EC.frame_to_be_available_and_switch_to_it((By.ID, "bvmf_iframe")))

    lclBtn1 = '/html/body/app-root/app-companies-home/div/div/div/div/div[2]/div[1]/div/div/a/i'
    lclBtnFind1 = driver.find_element('xpath', lclBtn1)
    # lclBtnFind1.click() # não deu certo
    driver.execute_script("arguments[0].click();", lclBtnFind1)

    # Adicionar espera para o elemento bdLink aparecer após o clique
    bdLink = '/html/body/app-root/app-companies-home/div/div/div/div/div[2]/div[2]/div/app-companies-home-filter-classification/form/div[2]/div[3]/div[2]/p/button'
    try:
        bd = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.XPATH, bdLink))
        )
        driver.execute_script("arguments[0].click();", bd)
    except Exception as e:
        print(f"Erro ao encontrar ou clicar no link de download: {e}")
        driver.quit()
        return None
    
    # necessário utilizar para dar tempo necessário para o download do arquivo.
    time.sleep(5) 

    driver.quit()

    # Drive da pasta download
    # o "r" se faz necessário porque as \ (barras) do windows são invertidas. Assim, ele funciona como uma exceção no uso do SO Windows. Os demais SO, não se faz necessário 
    pathDownload = r"C:\Users\Haroldo Melo\Downloads"
    # pathDownload = Path(__file__).parent
    # pathDownload = str(pathDownload).replace("/", "\\") # substitui as / (barras) do linux pelas \ (barras) do windows

    # arqZip = zipfile.ZipFile(pathDownload + r"\ClassifSetorial.zip")
    xlsx_path = pathDownload + r"\ClassifSetorial.xlsx"

    wb = load_workbook(filename=xlsx_path, read_only=False)
    ws = wb.active
    setores = pd.read_excel(xlsx_path, header=None)

    rowStart = None
    rowEnd = None  # Linha final da célula mesclada
    for merged_range in ws.merged_cells.ranges:
        rowS = merged_range.min_row - 1  # índice zero-based
        rowE = merged_range.max_row - 1    # índice zero-based

        if rowStart is None or rowS < rowStart:
            rowStart = rowS
            rowEnd = rowE
        elif rowS == rowStart:
            # Caso tenha mais de uma mesclagem iniciando na mesma linha, pega a maior final
            if rowE > rowEnd:
                rowEnd = rowE

    if rowStart is None:
        # Sem mesclagem, pega primeira linha não vazia
        # h1 = setores.dropna(how='all').index[0]
        rowStart = setores.index[0]
        rowEnd = setores.index[0]

    # Header até a última linha da mesclagem (inclusive) do cabeçalho
    header_df = setores.iloc[:rowEnd + 1, :]
    header_df = header_df.ffill(axis=0) # vertical/colunas | se fosse axis=1 seria horizontal/linhas
    header = []
    for col in header_df.columns:
        vals = header_df[col].dropna().astype(str).values
        if len(vals) > 1:
            merged_header = ' - '.join(vals)
        elif len(vals) == 1:
            merged_header = vals[0]
        else:
            merged_header = ''
        header.append(merged_header)
    # Função para remover texto antes de " - "
    header2 = []
    for col in header:
        if " - " in col:
            header2.append(col.split(" - ", 1)[1])
        else:
            header2.append(col)
        
    # data_row0 = rowEnd + 1
    # Ler dados a partir da linha seguinte ao final da mesclagem
    # define o dataframe a partir da linha final da mesclagem, primeira abaixo do cabeçalho
    df = pd.read_excel(xlsx_path, header=rowEnd)
    # redefinide cabeçalho para o dataframe que foi tratado acima
    df.columns = header2
    # print(f"Cabeçalho mesclado:\n{header2}")
    # print(f"Setores:\n{df}")

    # ind = []
    # headerDF = []
    # for index, col in enumerate(header2):
    #     if not str(col).startswith('Unnamed'):
    #         ind.append(index)
    #         headerDF.append(col)

    print('Colunas filtradas:')
    # df = df.dropna(how='all', axis=0) # opera nas linhas
    df = df.dropna(how='all', axis=1) # opera nas colunas
    df['SETOR'] = df['SETOR'].ffill()
    df['SUBSETOR'] = df['SUBSETOR'].ffill()
    df['SEGMENTO'] = df['SEGMENTO'].ffill()
    df['NOME DE PREGÃO'] = df['NOME DE PREGÃO'].ffill()
    df['CÓDIGO'] = df['CÓDIGO'].ffill()
    df['SEGMENTO DE NEGOCIAÇÃO'] = df['SEGMENTO DE NEGOCIAÇÃO'].ffill()
    # print(df)

    print(f"Filtragem por linha: {df.columns[1]}")
    # Alternativa: usar filter() do pandas
    # df = df.filter(regex=f'.*{df.columns[1]}*.', axis=0) # mantém nas linhas

    # Remove linhas onde a primeira coluna CONTÉM a palavra do cabeçalho
    # Exemplo: se o cabeçalho é "SETOR", remove linhas que tenham "O SETOR ECONOMICO"
    df = df[~df.iloc[:, 1].astype(str).str.contains(df.columns[1], case=False, na=False, regex=False)]

    # Removendo linhas duplicadas
    df = df.drop_duplicates()
    # remove linhas que tem valores n/a (null)
    df = df.dropna() # remove linhas que tem valores n/a (null)

    # df = df[~df.eq(df.columns).all(axis=1)]
    # df = df[~df.eq(df.columns).all(axis=0)]
    
    # setores.columns # apresente o nome das colunas
    # df.columns # apresente o nome das colunas
    # setores.columns = ['SUBSETOR', 'EMPRESA', 'TICKER'] # alterne o nome das colunas
    # df.columns = ['SUBSETOR', 'EMPRESA', 'TICKER'] # alterne o nome das colunas
        
    print(pathDownload + r"\ClassifSetorial.xlsx")
    os.remove(pathDownload + r"\ClassifSetorial.xlsx") # agora deu certo
    print(df)
    # print(diretorio)

    # Exportar dados tratados
    os.makedirs('B3/files/', exist_ok=True)  # Cria o diretório se não existir

    # Para CSV
    df.to_csv('B3/files/setores.csv', index=False, mode='w')  # mode='w' sobrescreve
    # Para Excel
    # df.to_excel('B3/files/setores.xlsx', index=False)  # Já sobrescreve por padrão

    # Para Parquet
    df.to_parquet('B3/files/setores.parquet', index=False)